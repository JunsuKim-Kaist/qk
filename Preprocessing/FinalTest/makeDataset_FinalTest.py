# create data and label for FER2013
# labels: 0=Angry, 1=Disgust, 2=Fear, 3=Happy, 4=Sad, 5=Surprise, 6=Neutral
import csv
import os
import numpy as np
import h5py
from random import shuffle
import cv2
import openpyxl
from os import listdir
from os.path import isfile, join
import argparse

parser = argparse.ArgumentParser(description='PyTorch Fer2013 CNN Training')

parser.add_argument('--input_path', type=str, default="/mnt/home/qualcomm/junsu/qk/FinalTestData/FaceAlignedVideo/", help='input data path')
parser.add_argument('--output_path', type=str, default="/mnt/home/qualcomm/junsu/qk/FinalTestData/TestDataset/", help='output data path')
parser.add_argument('--label_blind', '-lb', action='store_true', help='if labels of test data set is not accessible')
opt = parser.parse_args()

if not opt.label_blind:
    Emofile = '/mnt/home/qualcomm/customdata/customdata.xlsx'
else:
    # create Emofile
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"

    imageFileList = [f for f in listdir(opt.input_path) if isfile(join(opt.input_path, f))]
    imageFileNum = len(imageFileList)
    for index in range(imageFileNum):
        ws1["A" + str(index+1)] = imageFileList[index][3:-4] + ".mp4"
    wb.save(opt.output_path + 'FinalDataset.xlsx')
    Emofile = opt.output_path + 'FinalDataset.xlsx'

emoDict = {'anger': 0, 'disgust': 1, 'fear': 2, 'happy': 3, 'sad': 4, 'surprise': 5, 'neutral': 6}

# Creat the list to store the data and label information
customTest_x = []
customTest_y = []
customTest_titleIndex = []


videoPath = opt.input_path
datapath = opt.output_path
if not os.path.exists(os.path.dirname(datapath)):
    os.makedirs(os.path.dirname(datapath))


wb = openpyxl.load_workbook(Emofile)
# active sheet
ws = wb.get_sheet_by_name('Sheet1')
row_count = ws.max_row
col_count = ws.max_column
print(row_count)
print(col_count)

EmptyList = 0
for row in range(1, row_count+1):

    Training_type = "customTest"
    print("Training type is " + Training_type)

    if Training_type == "customTest":
        temp_list = []
        Title = ws.cell(row= row, column= 1).value
        if opt.label_blind:
            Label_int = -1
        else:
            Label_str = ws.cell(row= row, column= 2).value
            print(Label_str)
            Label_int = emoDict[Label_str]
        vidTitle = "fa_" + Title[:-4] + ".avi"
        print(vidTitle)
        cap = cv2.VideoCapture(videoPath + vidTitle)
        while cap.isOpened():
            ret, frame = cap.read()
            if ret:
                print("open")
                temp_list.append(frame)
                break
            else:
                break
        cap.release()
        '''
        if len(temp_list) > 3:
            print("over length")
            temp_list = temp_list[0:2]
        elif len(temp_list) < 3:
            print("less length")
            res = 3 - len(temp_list)
            if len(temp_list) == 0:
                EmptyList += 1
                continue
            print(len(temp_list))
            for i in range(res):
                temp_list.append(temp_list[len(temp_list)-1])
        else:
            pass
            '''
        I = np.asarray(temp_list)
        customTest_y.append(Label_int)
        customTest_x.append(I.tolist())
        customTest_titleIndex.append(row)


print(np.shape(customTest_x))
print(np.shape(customTest_y))

h5_datapath = os.path.join(opt.output_path, 'FinalData.h5')
datafile = h5py.File(h5_datapath, 'w')
datafile.create_dataset("FinalTest_pixel", dtype = 'uint8', data=customTest_x)
datafile.create_dataset("FinalTest_label", dtype = 'int64', data=customTest_y)
datafile.create_dataset("FinalTest_titleIndex", dtype = 'int64', data=customTest_titleIndex)
datafile.close()

print("Save data finish!!!")
print("Empty List Total : " + str(EmptyList))
