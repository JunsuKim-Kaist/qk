# create data and label for FER2013
# labels: 0=Angry, 1=Disgust, 2=Fear, 3=Happy, 4=Sad, 5=Surprise, 6=Neutral
import csv
import os
import numpy as np
import h5py
from random import shuffle
import cv2
import openpyxl

Emofile = '/mnt/home/qualcomm/customdata/customdata.xlsx'

emoDict = {'anger': 0, 'disgust': 1, 'fear': 2, 'happy': 3, 'sad': 4, 'surprise': 5, 'neutral': 6}

# Creat the list to store the data and label information
customTest_x = []
customTest_y = []
customTest_titleIndex = []

videoPath = '/mnt/home/qualcomm/customdata/alignedFaceVideo/'
datapath = '/mnt/home/qualcomm/customdata/'
if not os.path.exists(os.path.dirname(datapath)):
    os.makedirs(os.path.dirname(datapath))


wb = openpyxl.load_workbook(Emofile)
# active sheet
ws = wb.get_sheet_by_name('Sheet1')
row_count = ws.max_row
col_count = ws.max_column
print(row_count)
print(col_count)
randShuffleList = list(range(1, row_count+1))

EmptyList = 0
for row in range(1, row_count+1):
    
    Training_type = "customTest"
    print("Training type is " + Training_type)

    if Training_type == "customTest":
        temp_list = []
        Title = ws.cell(row= row, column= 1).value
        Label_str = ws.cell(row= row, column= 2).value
        print(Label_str)
        Label_int = emoDict[Label_str]
        vidTitle = "fa_" + Title[:-4] + ".avi"
        print(vidTitle)
        cap = cv2.VideoCapture(videoPath + vidTitle)
        while cap.isOpened():
            ret, frame = cap.read()
            if ret:
                print("open")
                temp_list.append(frame)
            else:
                break
        cap.release()
        if len(temp_list) > 10:
            print("over length")
            temp_list = temp_list[0:10]
        elif len(temp_list) < 10:
            print("less length")
            res = 10 - len(temp_list)
            if len(temp_list) == 0:
                EmptyList += 1
                continue
            print(len(temp_list))
            for i in range(res):
                temp_list.append(temp_list[len(temp_list)-1])
        else:
            pass
        I = np.asarray(temp_list)
        customTest_y.append(Label_int)
        customTest_x.append(I.tolist())
        customTest_titleIndex.append(row)

print(np.shape(customTest_x))
print(np.shape(customTest_y))

# customTest_asciititle = [index.encode("ascii", "ignore") for index in customTest_title]

h5_datapath = os.path.join(datapath, 'customdata.h5')
datafile = h5py.File(h5_datapath, 'w')
datafile.create_dataset("customTest_pixel", dtype = 'uint8', data=customTest_x)
datafile.create_dataset("customTest_label", dtype = 'int64', data=customTest_y)
datafile.create_dataset("customTest_titleIndex", dtype = 'int64', data=customTest_titleIndex)
datafile.close()

print("Save data finish!!!")
print("Empty List Total : " + str(EmptyList))
