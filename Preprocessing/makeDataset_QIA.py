# create data and label for FER2013
# labels: 0=Angry, 1=Disgust, 2=Fear, 3=Happy, 4=Sad, 5=Surprise, 6=Neutral
import csv
import os
import numpy as np
import h5py
from random import shuffle
import cv2
import openpyxl

Emofile = '/mnt/home/qualcomm/Hackathon2018/QIA-Hackathon 2018/Emotion Recognition/Dataset/multimodal_text.xlsx'

emoDict = {'anger': 0, 'disgust': 1, 'fear': 2, 'happy': 3, 'sad': 4, 'surprise': 5, 'neutral': 6}

# Creat the list to store the data and label information
Training_x = []
Training_y = []
PublicTest_x = []
PublicTest_y = []
PrivateTest_x = []
PrivateTest_y = []

videoPath = '/mnt/home/qualcomm/data/alignedFaceVideo/'
datapath = '/mnt/home/qualcomm/data/'
if not os.path.exists(os.path.dirname(datapath)):
    os.makedirs(os.path.dirname(datapath))


wb = openpyxl.load_workbook(Emofile)
# active sheet
ws = wb.get_sheet_by_name('Sheet1')
row_count = ws.max_row
col_count = ws.max_column
print(row_count)
print(col_count)
randShuffleList = list(range(1, row_count+1))
shuffle(randShuffleList)

EmptyList = 0
for row in range(1, row_count+1):
    
    Training_type = ""
    randomVal = randShuffleList[row - 1] / row_count
    print(str(row) + " / " + str(row_count) + " processing ... " )
    if randomVal < 0.8:
        Training_type = "Training"
    elif randomVal < 0.9:
        Training_type = "PublicTest"
    else:
        Training_type = "PrivateTest"

    print("Training type is " + Training_type)

    if Training_type == "Training":
        temp_list = []
        Title = ws.cell(row= row, column= 1).value
        Label_str = ws.cell(row= row, column= 2).value
        print(Label_str)
        Label_int = emoDict[Label_str]
        vidTitle = "fa_" + Title + ".avi"
        print(vidTitle)
        cap = cv2.VideoCapture(videoPath + vidTitle)
        while cap.isOpened():
            ret, frame = cap.read()
            if ret:
                temp_list.append(frame)
            else:
                break
        cap.release()
        if len(temp_list) > 10:
            print("over length")
            temp_list = temp_list[0:10]
        elif len(temp_list) < 10:
            print("less length")
            res = 10 - len(temp_list)
            if len(temp_list) == 0:
                EmptyList += 1
                continue
            print(len(temp_list))
            for i in range(res):
                temp_list.append(temp_list[len(temp_list)-1])
        else:
            pass
        I = np.asarray(temp_list)
        Training_y.append(Label_int)
        Training_x.append(I.tolist())

    if Training_type == "PublicTest":
        temp_list = []
        Title = ws.cell(row= row, column= 1).value
        Label_str = ws.cell(row= row, column= 2).value
        print(Label_str)
        Label_int = emoDict[Label_str]
        vidTitle = "fa_" + Title + ".avi"
        print(vidTitle)
        cap = cv2.VideoCapture(videoPath + vidTitle)
        while cap.isOpened():
            ret, frame = cap.read()
            if ret:
                temp_list.append(frame)
            else:
                break
        if len(temp_list) > 10:
            print("over length")
            temp_list = temp_list[0:10]
        elif len(temp_list) < 10:
            print(len(temp_list))
            print("less length")
            res = 10 - len(temp_list)
            if len(temp_list) == 0:
                EmptyList += 1
                continue
            for i in range(res):
                temp_list.append(temp_list[len(temp_list)-1])
        else:
            pass
        cap.release()

        I = np.asarray(temp_list)
        PublicTest_y.append(Label_int)
        PublicTest_x.append(I.tolist())

    if Training_type == "PrivateTest":
        temp_list = []
        Title = ws.cell(row= row, column= 1).value
        Label_str = ws.cell(row= row, column= 2).value
        print(Label_str)
        Label_int = emoDict[Label_str]
        vidTitle = "fa_" + Title + ".avi"
        print(vidTitle)
        cap = cv2.VideoCapture(videoPath + vidTitle)
        while cap.isOpened():
            ret, frame = cap.read()
            if ret:
                temp_list.append(frame)
            else:
                break
        cap.release()
        if len(temp_list) > 10:
            print("over length")
            temp_list = temp_list[0:10]
        elif len(temp_list) < 10:
            print("less length")
            print(len(temp_list))
            res = 10 - len(temp_list)
            if len(temp_list) == 0:
                EmptyList += 1
                continue
            for i in range(res):
                temp_list.append(temp_list[len(temp_list)-1])
        else:
            pass

        I = np.asarray(temp_list)

        PrivateTest_y.append(Label_int)
        PrivateTest_x.append(I.tolist())

print(np.shape(Training_x))
print(np.shape(PublicTest_x))
print(np.shape(PrivateTest_x))
h5_datapath = os.path.join(datapath, 'qia.h5')
datafile = h5py.File(h5_datapath, 'w')
datafile.create_dataset("Training_pixel", dtype = 'uint8', data=Training_x)
datafile.create_dataset("Training_label", dtype = 'int64', data=Training_y)
datafile.create_dataset("PublicTest_pixel", dtype = 'uint8', data=PublicTest_x)
datafile.create_dataset("PublicTest_label", dtype = 'int64', data=PublicTest_y)
datafile.create_dataset("PrivateTest_pixel", dtype = 'uint8', data=PrivateTest_x)
datafile.create_dataset("PrivateTest_label", dtype = 'int64', data=PrivateTest_y)
datafile.close()

print("Save data finish!!!")
print("Empty List Total : " + str(EmptyList))
