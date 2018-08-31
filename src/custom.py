''' Fer2013 Dataset class'''

from __future__ import print_function
from PIL import Image
import numpy as np
import h5py
import torch.utils.data as data
import openpyxl

class CUSTOM(data.Dataset):
    """`FER2013 Dataset.

    Args:
        train (bool, optional): If True, creates dataset from training set, otherwise
            creates from test set.
        transform (callable, optional): A function/transform that  takes in an PIL image
            and returns a transformed version. E.g, ``transforms.RandomCrop``
    """


    def __init__(self, split='Test', transform=None, h5file_path=None, Emofile_path=None):

        self.transform = transform
        self.split = split  # training set or test set
        self.data = h5py.File(h5file_path, 'r', driver='core')
        # now load the picked numpy arrays
        #for key in self.data.keys():
        #    print(key)
        #    print(type(self.data[key][0]))

        Emofile = Emofile_path
        wb = openpyxl.load_workbook(Emofile)
        ws = wb.get_sheet_by_name('Sheet1')
        datasetSize = ws.max_row
        #print(self.data['customTest_titleIndex'][0])
        self.title = []
        for index in range(datasetSize):
            rowIndex = self.data['FinalTest_titleIndex'][index]
            #print(rowIndex)
            #print(ws.cell(row= rowIndex, column=1).value)
            self.title.append(ws.cell(row= rowIndex, column=1).value)
        self.labels = self.data['FinalTest_label']
        self.data = self.data['FinalTest_pixel']

        self.data = np.asarray(self.data)
        self.data = self.data.reshape((datasetSize, 1, 48, 48, 3))

        #print(self.title)

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (image, target) where target is index of the target class.
        """
        img, target = self.data[index], self.labels[index]
        title = self.title[index]
        #print(title)
        imgList = np.empty((3, 44, 44, 0), int)
        concat_axis = 3

        # doing this so that it is consistent with all other datasets
        # to return a PIL Image

        #print(img.shape)

        tuple_zero = (0,)
        # imgList = np.empty(tuple_zero + img.shape[1:], int)
        #imgList = np.empty((0, 44, 44, 3), int)
        #print(img.shape)
        #print(imgList.shape)
        for frame in range(img.shape[0]):
            if frame == 0:
                imgFrame = Image.fromarray(img[frame,:,:,:])
                #print(imgFrame.size)
                if self.transform is not None:
                    imgFrame = self.transform(imgFrame)
                image_np = np.array(imgFrame)

                return title, image_np, target


    def __len__(self):
        return len(self.data)
